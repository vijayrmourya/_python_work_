# Reading existing values in excel file
# writing the col_values array values in place or values matched in 'value' array in excel
# import library to edit excel
import openpyxl

# give the sheet path
path = "G:\Z_Python\Book1.xlsx"
# creating sheet object to edit the excel
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
# give desired column_name name and values
column_name = "Rte_VehStatus_In_100_V_x_RCarHMIManagementSensorInput_V_x_RCarHMIManagementSensorInput.V_x_MasterObjDataModified[0].V_x_ObjectType"
col_values = [0, 1]
value = ['zero', 'two']
flag = 0
# getting max number of column in the sheet
max_col = sheet_obj.max_column
# for loop to find the desired column in sheet
for i in range(1, max_col + 1):
    # always change the row value based on your sheets column name position
    # if the column name is in 5th row give row=5
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value)
    # if desired column is found then starting the row editing
    if column_name == cell_obj.value:
        flag = 1
        # getting max number of column in the sheet
        m_row = sheet_obj.max_row
        # row editing for loop
        for j in range(1, m_row + 1):
            cell_obj = sheet_obj.cell(row=j, column=i)
            print(cell_obj.value)
            n = cell_obj.value
            for k in range(len(value)):
                if n == value[k]:
                    print('col_values matched')
                    c1 = sheet_obj.cell(row=j, column=i)
                    c1.value = col_values[k]
                    print(c1.value)
                    wb_obj.save(path)
                    break
    if (flag == 1):
        break
