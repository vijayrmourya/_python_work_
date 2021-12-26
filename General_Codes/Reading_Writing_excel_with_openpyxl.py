# Reading existing values in excel file
# writing the case array values in place or values matched in 'value' array in excel
# import library to edit excel
import openpyxl

# give the sheet path
path = "G:\Z_Python\Book1.xlsx"
# creating sheet object to edit the excel
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
# give desired signal name and values
signal = "V_x_ST_FRONT_CAM_OBJ[0]"
case = [5, 6, 7, 8, 9, 10]
value = [10, 11, 12, 13, 14, 15]
flag = 0
# getting max number of column in the sheet
max_col = sheet_obj.max_column
# for loop to find the desired column in sheet
for i in range(1, max_col + 1):
    # always change the row value based on your sheets column name position
    # if the column name is in 5th row give row=5
    cell_obj = sheet_obj.cell(row=1, column=i)
    # print(cell_obj.value)
    # if desired column is found then starting the row editing
    if signal == cell_obj.value:
        flag = 1
        # getting max number of column in the sheet
        m_row = sheet_obj.max_row
        # row editing for loop
        for j in range(1, m_row + 1):
            cell_obj = sheet_obj.cell(row=j, column=i)
            # print(cell_obj.value)
            n = cell_obj.value
            for k in range(len(value)):
                if n == value[k]:
                    print('case matched')
                    c1 = sheet_obj.cell(row=j, column=i)
                    c1.value = case[k]
                    # print(c1.value)
                    wb_obj.save("G:\Z_Python\Book1.xlsx")
                    break
    if (flag == 1):
        break
