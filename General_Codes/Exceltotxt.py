import openpyxl

xl_path = "G:\\--Entire_Github--\\Python\\My_Python_Work\\Sample_Employee_data_xls.xlsx"
txt_file = "G:\\--Entire_Github--\\Python\\My_Python_Work\\readme.txt"
wb_obj = openpyxl.load_workbook(xl_path)
sheet_obj = wb_obj.active
max_cols = sheet_obj.max_column
max_rows = sheet_obj.max_row
print("started writing!")
with open(txt_file, 'w') as f:
    for i in range(1, max_rows + 1):
        for j in range(1, max_cols + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            # print(cell_obj.value)
            content = cell_obj.value
            f.write(str(cell_obj.value) + "|")
        f.write("\n")
f.close()
print(f"finished! you may find your file at {txt_file}")
